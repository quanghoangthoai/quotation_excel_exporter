import frappe
import io
import os
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, numbers
from datetime import datetime
import tempfile

@frappe.whitelist()
def export_excel_api(quotation_name):
    # Validate input
    if not quotation_name or not frappe.db.exists("Quotation", quotation_name):
        frappe.throw("Invalid or non-existent Quotation")

    # Fetch Quotation and Customer
    quotation = frappe.get_doc("Quotation", quotation_name)
    customer = frappe.get_doc("Customer", quotation.party_name) if quotation.party_name else None

    # Fetch Company Details
    company = frappe.get_doc("Company", frappe.defaults.get_user_default("company"))
    # Fetch company address via Dynamic Link
    address_name = frappe.db.get_value(
        "Dynamic Link",
        {"link_doctype": "Company", "link_name": company.name, "parenttype": "Address"},
        "parent"
    )
    address = "Your Address"
    if address_name:
        addr = frappe.get_doc("Address", address_name)
        address = addr.address_line1 or ""
        if addr.address_line2:
            address += ", " + addr.address_line2
        if addr.city:
            address += ", " + addr.city
        if addr.country:
            address += ", " + addr.country

    company_details = {
        "name": company.company_name or "Your Company Name",
        "address": address,
        "phone_no": company.phone_no or "Your Phone Number",
        "website": company.website or "Your Website",
        "logo": company.get("company_logo")  # Use company_logo field directly, no fallback
    }

    # Initialize Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo giá"

    # Define Styles
    font_13 = Font(name="Times New Roman", size=13)
    font_13_bold = Font(name="Times New Roman", size=13, bold=True)
    font_16 = Font(name="Times New Roman", size=16, bold=True)
    font_18 = Font(name="Times New Roman", size=18, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid")
    currency_format = '#,##0_₫'

    # Logo (Optional)
    logo_path = company_details["logo"]
    if logo_path and os.path.exists(logo_path):
        try:
            logo_img = XLImage(logo_path)
            logo_img.width = 202
            logo_img.height = 60
            ws.add_image(logo_img, "A1")
            ws.merge_cells("A1:B3")
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 20
        except Exception as e:
            frappe.log_error(
                message=f"Failed to add logo: {e}",
                title="Excel Exporter Logo Error"
            )

    # Company Details
    ws.merge_cells("D2:H3")
    ws["D2"] = company_details["name"]
    ws["D2"].font = font_18
    ws["D2"].alignment = center_alignment

    ws.merge_cells("A5:B5")
    ws["A5"] = "Địa chỉ:"
    ws.merge_cells("C5:N5")
    ws["C5"] = company_details["address"]
    ws.merge_cells("A6:B6")
    ws["A6"] = "Hotline:"
    ws.merge_cells("C6:N6")
    ws["C6"] = company_details["phone_no"]
    ws.merge_cells("A7:B7")
    ws["A7"] = "Website:"
    ws.merge_cells("C7:N7")
    ws["C7"] = company_details["website"]

    for cell in ["A5", "A6", "A7", "C5", "C6", "C7"]:
        ws[cell].font = font_13
        ws[cell].alignment = left_alignment

    # Title
    ws.merge_cells("A9:N9")
    ws["A9"] = "PHIẾU BÁO GIÁ BÁN HÀNG"
    ws["A9"].font = font_16
    ws["A9"].alignment = center_alignment

    # Introduction
    ws.merge_cells("A11:N11")
    ws["A11"] = "Lời đầu tiên, xin cảm ơn Quý khách hàng đã quan tâm đến sản phẩm nội thất của công ty chúng tôi."
    ws["A11"].font = font_13
    ws["A11"].alignment = left_alignment
    ws.merge_cells("A12:N12")
    ws["A12"] = "Chúng tôi xin gửi đến Quý khách hàng Bảng báo giá như sau:"
    ws["A12"].font = font_13
    ws["A12"].alignment = left_alignment

    # Customer Details
    ws["A13"] = "Khách hàng:"
    ws["A13"].font = font_13_bold
    ws.merge_cells("B13:H13")
    ws["B13"] = customer.customer_name if customer else ""
    ws["I13"] = "Điện thoại:"
    ws.merge_cells("J13:N13")
    phone = ""
    contact_name = frappe.db.get_value(
        "Dynamic Link",
        {"link_doctype": "Customer", "link_name": quotation.party_name, "parenttype": "Contact"},
        "parent"
    )
    if contact_name:
        contact = frappe.get_doc("Contact", contact_name)
        phone = contact.mobile_no or contact.phone or ""
    ws["J13"] = phone

    ws["A14"] = "Địa chỉ:"
    ws.merge_cells("B14:N14")
    address = ""
    address_name = frappe.db.get_value(
        "Dynamic Link",
        {"link_doctype": "Customer", "link_name": quotation.party_name, "parenttype": "Address"},
        "parent"
    )
    if address_name:
        addr = frappe.get_doc("Address", address_name)
        address = addr.address_line1 or ""
        if addr.address_line2:
            address += ", " + addr.address_line2
        if addr.city:
            address += ", " + addr.city
        if addr.country:
            address += ", " + addr.country
    ws["B14"] = address

    for cell in ["A13", "B13", "I13", "J13", "A14", "B14"]:
        ws[cell].font = font_13 if cell != "A13" else font_13_bold
        ws[cell].alignment = left_alignment

    # Table Headers
    headers = [
        "STT", "Tên sản phẩm", "", "", "Kích thước sản phẩm", "", "Mã hàng",
        "SL", "Hình ảnh", "", "Đơn vị", "Đơn giá", "CK (%)", "Thành tiền"
    ]
    header_row = 16
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = font_13_bold
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    ws.merge_cells(f"B{header_row}:D{header_row}")
    ws.merge_cells(f"E{header_row}:F{header_row}")
    ws.merge_cells(f"I{header_row}:J{header_row}")

    # Table Data
    temp_files = []
    for idx, item in enumerate(quotation.items, 1):
        row = header_row + idx
        ws.cell(row=row, column=1, value=idx).font = font_13
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=item.item_name or "").font = font_13
        ws.merge_cells(f"E{row}:F{row}")
        ws.cell(row=row, column=5, value=frappe.db.get_value("Quotation Item", item.name, "size") or "").font = font_13
        ws.cell(row=row, column=7, value=item.item_code or "").font = font_13
        ws.cell(row=row, column=8, value=item.qty or 0).font = font_13
        ws.cell(row=row, column=11, value="Bộ").font = font_13
        rate_cell = ws.cell(row=row, column=12)
        rate_cell.value = item.price_list_rate or 0
        rate_cell.number_format = currency_format
        rate_cell.font = font_13
        ws.cell(row=row, column=13, value=item.discount_percentage or 0).font = font_13
        amt_cell = ws.cell(row=row, column=14)
        amt_cell.value = f"=L{row}*H{row}*(1-M{row}/100)"
        amt_cell.number_format = currency_format
        amt_cell.font = font_13

        # Image Handling
        if item.image:
            try:
                image_path = None
                if item.image.startswith("/files/"):
                    image_path = frappe.get_site_path("public", item.image.lstrip("/"))
                elif item.image.startswith("http"):
                    resp = requests.get(item.image, timeout=5)
                    if resp.status_code == 200:
                        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                        tmp.write(resp.content)
                        tmp.close()
                        image_path = tmp.name
                        temp_files.append(image_path)
                if image_path and os.path.exists(image_path):
                    img = XLImage(image_path)
                    img.width = 140
                    img.height = 90
                    ws.merge_cells(f"I{row}:J{row}")  # Merge I and J columns for the image
                    ws.add_image(img, f"I{row}")
                    ws.row_dimensions[row].height = 100
            except Exception as e:
                err_msg = f"Failed to add image for item {item.item_code}: {e}"[:100]
                frappe.log_error(
                    message=err_msg,
                    title=f"Excel Exporter Image Error [{item.item_code}]"
                )

        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment

    # Totals
    start_data_row = header_row + 1
    end_data_row = header_row + len(quotation.items)
    total_row = end_data_row + 1

    ws.cell(row=total_row, column=1, value="A").font = font_13
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=13)
    ws.cell(row=total_row, column=2, value="Tổng cộng").font = font_13
    total_cell = ws.cell(row=total_row, column=14)
    total_cell.value = f"=SUM(N{start_data_row}:N{end_data_row})"
    total_cell.number_format = currency_format
    total_cell.font = font_13
    for col in range(1, 15):
        ws.cell(row=total_row, column=col).border = thin_border

    add_row = total_row + 1
    ws.cell(row=add_row, column=1, value="B").font = font_13
    ws.merge_cells(start_row=add_row, start_column=2, end_row=add_row, end_column=13)
    ws.cell(row=add_row, column=2, value="Phụ phí").font = font_13
    add_cell = ws.cell(row=add_row, column=14)
    add_cell.value = 0
    add_cell.number_format = currency_format
    add_cell.font = font_13
    for col in range(1, 15):
        ws.cell(row=add_row, column=col).border = thin_border

    paid_row = add_row + 1
    ws.cell(row=paid_row, column=1, value="C").font = font_13
    ws.merge_cells(start_row=paid_row, start_column=2, end_row=paid_row, end_column=13)
    ws.cell(row=paid_row, column=2, value="Đã thanh toán").font = font_13
    paid_amt_cell = ws.cell(row=paid_row, column=14)
    paid_amt_cell.value = 0
    paid_amt_cell.number_format = currency_format
    paid_amt_cell.font = font_13
    for col in range(1, 15):
        ws.cell(row=paid_row, column=col).border = thin_border

    final_row = paid_row + 1
    ws.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=13)
    ws.cell(row=final_row, column=1, value="Tổng tiền thanh toán (A+B-C)").font = font_13
    final_cell = ws.cell(row=final_row, column=14)
    final_cell.value = f"=N{total_row}+N{add_row}-N{paid_row}"
    final_cell.number_format = currency_format
    final_cell.font = font_13
    for col in range(1, 15):
        ws.cell(row=final_row, column=col).border = thin_border

    for r in range(total_row, final_row + 1):
        ws.cell(row=r, column=1).alignment = left_alignment if r == final_row else center_alignment
        ws.cell(row=r, column=2).alignment = left_alignment
        ws.cell(row=r, column=14).alignment = center_alignment

    # Footer
    footer_start = final_row + 2
    ws.merge_cells(start_row=footer_start, start_column=1, end_row=footer_start, end_column=5)
    ws.cell(row=footer_start, column=1, value="Khách hàng").font = font_13_bold
    ws.cell(row=footer_start, column=1).alignment = center_alignment
    ws.merge_cells(start_row=footer_start, start_column=6, end_row=footer_start, end_column=10)
    ws.cell(row=footer_start, column=6, value="Người giao hàng").font = font_13_bold
    ws.cell(row=footer_start, column=6).alignment = center_alignment
    ws.merge_cells(start_row=footer_start, start_column=11, end_row=footer_start, end_column=14)
    date = quotation.transaction_date or datetime.now()
    ws.cell(row=footer_start, column=11, value=f"Ngày {date.day:02d} Tháng {date.month:02d} Năm {date.year}").font = font_13_bold
    ws.cell(row=footer_start, column=11).alignment = center_alignment

    ws.merge_cells(start_row=footer_start + 1, start_column=1, end_row=footer_start + 1, end_column=5)
    ws.cell(row=footer_start + 1, column=1, value="(Ký và ghi rõ họ tên)").font = font_13
    ws.cell(row=footer_start + 1, column=1).alignment = center_alignment
    ws.merge_cells(start_row=footer_start + 1, start_column=6, end_row=footer_start + 1, end_column=10)
    ws.cell(row=footer_start + 1, column=6, value="(Ký và ghi rõ họ tên)").font = font_13
    ws.cell(row=footer_start + 1, column=6).alignment = center_alignment
    ws.merge_cells(start_row=footer_start + 1, start_column=11, end_row=footer_start + 1, end_column=14)
    ws.cell(row=footer_start + 1, column=11, value="(Ký và ghi rõ họ tên)").font = font_13
    ws.cell(row=footer_start + 1, column=11).alignment = center_alignment

    # Notes
    remarks_row = footer_start + 3
    ws.merge_cells(start_row=remarks_row, start_column=1, end_row=remarks_row, end_column=14)
    ws.cell(row=remarks_row, column=1, value="Lưu ý: Không đổi trả sản phẩm mẫu trừ trường hợp sản phẩm bị lỗi từ nhà sản xuất").font = font_13_bold
    ws.cell(row=remarks_row, column=1).alignment = left_alignment
    ws.merge_cells(start_row=remarks_row + 1, start_column=1, end_row=remarks_row + 1, end_column=14)
    ws.cell(row=remarks_row + 1, column=1, value="Hình thức thanh toán:").font = font_13_bold
    ws.cell(row=remarks_row + 1, column=1).alignment = left_alignment
    ws.merge_cells(start_row=remarks_row + 2, start_column=1, end_row=remarks_row + 2, end_column=14)
    ws.cell(row=remarks_row + 2, column=1, value="- Thanh toán 100% giá trị đơn hàng khi nhận được hàng").font = font_13
    ws.cell(row=remarks_row + 2, column=1).alignment = left_alignment
    ws.merge_cells(start_row=remarks_row + 3, start_column=1, end_row=remarks_row + 3, end_column=14)
    ws.cell(row=remarks_row + 3, column=1, value="- Đặt hàng đặt cọc trước 30% giá trị đơn hàng").font = font_13
    ws.cell(row=remarks_row + 3, column=1).alignment = left_alignment

    # Column Widths
    column_widths = {
        "A": 12, "B": 20, "C": 5, "D": 5,
        "E": 15, "F": 5, "G": 15, "H": 8,
        "I": 15, "J": 8, "K": 10, "L": 20,
        "M": 10, "N": 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save and Output
    try:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        frappe.local.response.filename = f"Bao_gia_{quotation.name}.xlsx"
        frappe.local.response.filecontent = output.read()
        frappe.local.response.type = "binary"
    except Exception as e:
        frappe.throw(f"Failed to generate Excel file: {e}")
    finally:
        for temp_file in temp_files:
            try:
                os.unlink(temp_file)
            except Exception:
                pass
